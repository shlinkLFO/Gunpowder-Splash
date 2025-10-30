from fastapi import APIRouter, HTTPException, Body
from typing import Optional, List
from pydantic import BaseModel
import openpyxl
import xlrd
from pathlib import Path

from app.services import file_service

router = APIRouter(prefix="/api/files", tags=["files"])


class FileCreate(BaseModel):
    path: str
    content: str = ""


class FileUpdate(BaseModel):
    content: str


class FolderCreate(BaseModel):
    path: str


class FileMove(BaseModel):
    source: str
    target_folder: str


class ExcelUpdate(BaseModel):
    data: List[List[str]]


@router.get("/tree")
async def get_file_tree():
    """Get hierarchical file tree structure"""
    tree = file_service.build_file_tree()
    return {"tree": tree}


@router.get("/info/{file_path:path}")
async def get_file_info(file_path: str):
    """Get file information"""
    info = file_service.get_file_info(file_path)
    if info is None:
        raise HTTPException(status_code=404, detail="File not found")
    return info


@router.get("/{file_path:path}")
async def read_file(file_path: str):
    """Read a file from workspace"""
    content = file_service.read_file(file_path)
    if content is None:
        raise HTTPException(status_code=404, detail="File not found")
    return {"content": content, "path": file_path}


@router.post("/create")
async def create_file(data: FileCreate):
    """Create a new file"""
    success = file_service.create_file(data.path, data.content)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to create file")
    return {"success": True, "path": data.path}


@router.post("/folder")
async def create_folder(data: FolderCreate):
    """Create a new folder"""
    success = file_service.create_folder(data.path)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to create folder")
    return {"success": True, "path": data.path}


@router.put("/{file_path:path}")
async def update_file(file_path: str, data: FileUpdate):
    """Update an existing file"""
    success = file_service.update_file(file_path, data.content)
    if not success:
        raise HTTPException(status_code=404, detail="File not found or update failed")
    return {"success": True, "path": file_path}


@router.delete("/{file_path:path}")
async def delete_file(file_path: str):
    """Delete a file or folder"""
    success = file_service.delete_file_or_folder(file_path)
    if not success:
        raise HTTPException(status_code=404, detail="File or folder not found")
    return {"success": True, "path": file_path}


@router.post("/move")
async def move_file(data: FileMove):
    """Move a file to a different folder"""
    new_path = file_service.move_file(data.source, data.target_folder)
    if new_path is None:
        raise HTTPException(status_code=404, detail="File not found or move failed")
    return {"success": True, "new_path": new_path}


@router.get("/parse-excel/{file_path:path}")
async def parse_excel_file(file_path: str):
    """Parse Excel file (.xlsx, .xls) and return tabular data"""
    try:
        from app.services.file_service import WORKSPACE_DIR
        workspace_resolved = WORKSPACE_DIR.resolve()
        full_path = (WORKSPACE_DIR / file_path).resolve()
        
        # Security: Ensure the resolved path is within workspace directory
        try:
            full_path.relative_to(workspace_resolved)
        except ValueError:
            raise HTTPException(status_code=403, detail="Access denied: path outside workspace")
        
        if not full_path.exists() or not full_path.is_file():
            raise HTTPException(status_code=404, detail="File not found")
        
        file_ext = full_path.suffix.lower()
        
        if file_ext == '.xlsx':
            # Parse .xlsx files using openpyxl
            wb = openpyxl.load_workbook(full_path, data_only=True)
            ws = wb.active
            
            data = []
            for row in ws.iter_rows(values_only=True):
                # Convert row tuple to list and handle None values
                data.append([str(cell) if cell is not None else '' for cell in row])
            
            return {
                "data": data,
                "path": file_path,
                "sheet_name": ws.title,
                "file_type": "xlsx"
            }
        
        elif file_ext == '.xls':
            # Parse .xls files using xlrd
            wb = xlrd.open_workbook(full_path)
            ws = wb.sheet_by_index(0)
            
            data = []
            for row_idx in range(ws.nrows):
                row = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    row.append(str(cell.value) if cell.value is not None else '')
                data.append(row)
            
            return {
                "data": data,
                "path": file_path,
                "sheet_name": ws.name,
                "file_type": "xls"
            }
        
        else:
            raise HTTPException(status_code=400, detail="File must be .xlsx or .xls format")
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error parsing Excel file: {str(e)}")


@router.put("/save-excel/{file_path:path}")
async def save_excel_file(file_path: str, data: ExcelUpdate):
    """Save edited data to Excel file (.xlsx, .xls)"""
    try:
        from app.services.file_service import WORKSPACE_DIR
        workspace_resolved = WORKSPACE_DIR.resolve()
        full_path = (WORKSPACE_DIR / file_path).resolve()
        
        # Security: Ensure the resolved path is within workspace directory
        try:
            full_path.relative_to(workspace_resolved)
        except ValueError:
            raise HTTPException(status_code=403, detail="Access denied: path outside workspace")
        
        if not full_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        
        file_ext = full_path.suffix.lower()
        
        if file_ext == '.xlsx':
            # Load existing workbook to preserve other sheets, formulas, and formatting
            wb = openpyxl.load_workbook(full_path)
            ws = wb.active
            
            # Clear existing data in active sheet
            ws.delete_rows(1, ws.max_row)
            
            # Write new data
            for row_data in data.data:
                ws.append(row_data)
            
            wb.save(full_path)
            return {"success": True, "path": file_path}
        
        elif file_ext == '.xls':
            # For .xls, convert to .xlsx since xlrd doesn't support writing
            # We'll save as .xlsx with a note
            raise HTTPException(
                status_code=400, 
                detail=".xls format doesn't support editing. Please convert to .xlsx first."
            )
        
        else:
            raise HTTPException(status_code=400, detail="File must be .xlsx or .xls format")
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving Excel file: {str(e)}")
